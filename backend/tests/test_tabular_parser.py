"""Tests for tabular file parsing utilities."""

import datetime
from io import BytesIO

import openpyxl
import pytest

from app.utils.tabular_parser import detect_columns_csv, detect_columns_xlsx, parse_csv, parse_xlsx


class TestParseCsv:
    """Tests for parse_csv function."""

    def test_parse_csv_utf8_bom(self, test_db):
        content = b"\xef\xbb\xbfcode;name;city\n001;Acme Corp;Paris\n002;Beta Ltd;London\n"
        rows = parse_csv(content)
        assert len(rows) == 2
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "Acme Corp"
        assert rows[0]["city"] == "Paris"

    def test_parse_csv_semicolon_delimiter(self, test_db):
        content = b"code;name;city\n001;Acme Corp;Paris\n"
        rows = parse_csv(content)
        assert len(rows) == 1
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "Acme Corp"

    def test_parse_csv_whitespace_trimming(self, test_db):
        content = b"code;name;city\n  001  ;  Acme Corp  ;  Paris  \n"
        rows = parse_csv(content)
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "Acme Corp"
        assert rows[0]["city"] == "Paris"

    def test_parse_csv_windows_1252_fallback(self, test_db):
        content = b"code;name;city\n001;Caf\xe9 Corp;Paris\n"
        rows = parse_csv(content)
        assert len(rows) == 1
        assert rows[0]["name"] == "Café Corp"

    def test_parse_csv_empty_file(self, test_db):
        rows = parse_csv(b"")
        assert rows == []

    def test_parse_csv_quoted_values_with_semicolons(self, test_db):
        content = b'code;name;address\n001;"Acme; Corp";"123; Main St"\n'
        rows = parse_csv(content)
        assert len(rows) == 1
        assert rows[0]["name"] == "Acme; Corp"
        assert rows[0]["address"] == "123; Main St"


class TestDetectColumnsCsv:
    """Tests for detect_columns_csv function."""

    def test_detect_columns_returns_headers(self, test_db):
        content = b"code;name;city;country\n001;Acme;Paris;France\n"
        columns = detect_columns_csv(content)
        assert columns == ["code", "name", "city", "country"]

    def test_detect_columns_with_bom(self, test_db):
        content = b"\xef\xbb\xbfcode;name;city\n001;Acme;Paris\n"
        columns = detect_columns_csv(content)
        assert columns == ["code", "name", "city"]

    def test_detect_columns_trims_whitespace(self, test_db):
        content = b"  code  ;  name  ;  city  \n001;Acme;Paris\n"
        columns = detect_columns_csv(content)
        assert columns == ["code", "name", "city"]

    def test_detect_columns_empty_file(self, test_db):
        columns = detect_columns_csv(b"")
        assert columns == []


def _make_xlsx(rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    """Build an in-memory xlsx workbook with the given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseXlsx:
    """Tests for parse_xlsx function."""

    def test_parse_xlsx_strings(self, test_db):
        content = _make_xlsx([["code", "name", "city"], ["001", "Acme Corp", "Paris"]])
        rows = parse_xlsx(content)
        assert rows == [{"code": "001", "name": "Acme Corp", "city": "Paris"}]

    def test_parse_xlsx_preserves_numbers(self, test_db):
        content = _make_xlsx([["code", "amount", "qty"], ["001", 1234.56, 7]])
        rows = parse_xlsx(content)
        assert rows[0]["amount"] == 1234.56
        assert isinstance(rows[0]["amount"], float)
        assert rows[0]["qty"] == 7
        assert isinstance(rows[0]["qty"], int)

    def test_parse_xlsx_preserves_booleans(self, test_db):
        content = _make_xlsx([["code", "active"], ["001", True], ["002", False]])
        rows = parse_xlsx(content)
        assert rows[0]["active"] is True
        assert rows[1]["active"] is False

    def test_parse_xlsx_dates_iso(self, test_db):
        content = _make_xlsx(
            [
                ["code", "created"],
                ["001", datetime.datetime(2025, 3, 1, 0, 0, 0)],
                ["002", datetime.datetime(2025, 3, 2, 14, 30, 0)],
            ]
        )
        rows = parse_xlsx(content)
        assert rows[0]["created"] == "2025-03-01"
        assert rows[1]["created"] == "2025-03-02T14:30:00"

    def test_parse_xlsx_empty_cells_are_none(self, test_db):
        content = _make_xlsx([["code", "name", "city"], ["001", None, "Paris"]])
        rows = parse_xlsx(content)
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] is None
        assert rows[0]["city"] == "Paris"

    def test_parse_xlsx_trims_string_values(self, test_db):
        content = _make_xlsx([["code", "name"], ["  001  ", "  Acme  "]])
        rows = parse_xlsx(content)
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "Acme"

    def test_parse_xlsx_drops_trailing_empty_rows(self, test_db):
        content = _make_xlsx(
            [
                ["code", "name"],
                ["001", "Acme"],
                [None, None],
                [None, None],
            ]
        )
        rows = parse_xlsx(content)
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_xlsx_keeps_rows_with_some_none(self, test_db):
        content = _make_xlsx(
            [
                ["code", "name", "city"],
                ["001", None, "Paris"],
                [None, "Beta", None],
            ]
        )
        rows = parse_xlsx(content)
        assert len(rows) == 2
        assert rows[0] == {"code": "001", "name": None, "city": "Paris"}
        assert rows[1] == {"code": None, "name": "Beta", "city": None}

    def test_parse_xlsx_reads_first_sheet_only(self, test_db):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1.append(["code", "name"])
        ws1.append(["001", "Acme"])
        ws2 = wb.create_sheet("Second")
        ws2.append(["code", "name"])
        ws2.append(["999", "Ignored"])
        buf = BytesIO()
        wb.save(buf)

        rows = parse_xlsx(buf.getvalue())
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_xlsx_formula_cached_value(self, test_db):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["a", "b", "sum"])
        ws.append([2, 3, "=A2+B2"])
        ws["C2"].value = 5
        buf = BytesIO()
        wb.save(buf)

        rows = parse_xlsx(buf.getvalue())
        assert rows[0]["sum"] == 5

    def test_parse_xlsx_empty_file(self, test_db):
        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        rows = parse_xlsx(buf.getvalue())
        assert rows == []

    def test_parse_xlsx_invalid_bytes_raises(self, test_db):
        with pytest.raises(ValueError, match="Could not read Excel file"):
            parse_xlsx(b"not a real xlsx file")


class TestDetectColumnsXlsx:
    """Tests for detect_columns_xlsx function."""

    def test_detect_columns_xlsx_returns_headers(self, test_db):
        content = _make_xlsx([["code", "name", "city"], ["001", "Acme", "Paris"]])
        columns = detect_columns_xlsx(content)
        assert columns == ["code", "name", "city"]

    def test_detect_columns_xlsx_trims_whitespace(self, test_db):
        content = _make_xlsx([["  code  ", "  name  "], ["001", "Acme"]])
        columns = detect_columns_xlsx(content)
        assert columns == ["code", "name"]

    def test_detect_columns_xlsx_filters_blank_header_cells(self, test_db):
        content = _make_xlsx([["code", None, "city", ""], ["001", "x", "Paris", "y"]])
        columns = detect_columns_xlsx(content)
        assert columns == ["code", "city"]

    def test_detect_columns_xlsx_empty_workbook(self, test_db):
        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        columns = detect_columns_xlsx(buf.getvalue())
        assert columns == []


class TestParseFileDispatch:
    """Tests for the parse_file dispatch helper."""

    def test_parse_file_dispatches_csv(self, test_db):
        from app.utils.tabular_parser import parse_file

        content = b"code;name\n001;Acme\n"
        rows = parse_file(content, "vendors.csv")
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_file_dispatches_csv_with_custom_delimiter(self, test_db):
        from app.utils.tabular_parser import parse_file

        content = b"code,name\n001,Acme\n"
        rows = parse_file(content, "vendors.csv", delimiter=",")
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_file_dispatches_xlsx(self, test_db):
        from app.utils.tabular_parser import parse_file

        content = _make_xlsx([["code", "name"], ["001", "Acme"]])
        rows = parse_file(content, "vendors.xlsx")
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_file_extension_case_insensitive(self, test_db):
        from app.utils.tabular_parser import parse_file

        content = _make_xlsx([["code", "name"], ["001", "Acme"]])
        rows = parse_file(content, "VENDORS.XLSX")
        assert rows == [{"code": "001", "name": "Acme"}]

    def test_parse_file_unknown_extension_raises(self, test_db):
        import pytest

        from app.utils.tabular_parser import parse_file

        with pytest.raises(ValueError, match="Unsupported file format"):
            parse_file(b"junk", "vendors.txt")

    def test_parse_file_xls_legacy_rejected(self, test_db):
        import pytest

        from app.utils.tabular_parser import parse_file

        with pytest.raises(ValueError, match="Unsupported file format"):
            parse_file(b"junk", "vendors.xls")


class TestDetectHeadersDispatch:
    """Tests for the detect_headers dispatch helper."""

    def test_detect_headers_csv_semicolon_sniff(self, test_db):
        from app.utils.tabular_parser import detect_headers

        columns, delimiter = detect_headers(b"code;name;city\n001;Acme;Paris\n", "vendors.csv")
        assert columns == ["code", "name", "city"]
        assert delimiter == ";"

    def test_detect_headers_csv_comma_sniff(self, test_db):
        from app.utils.tabular_parser import detect_headers

        columns, delimiter = detect_headers(b"code,name,city\n001,Acme,Paris\n", "vendors.csv")
        assert columns == ["code", "name", "city"]
        assert delimiter == ","

    def test_detect_headers_tsv_rejected(self, test_db):
        import pytest

        from app.utils.tabular_parser import detect_headers

        with pytest.raises(ValueError, match="Unsupported file format"):
            detect_headers(b"code\tname\tcity\n", "vendors.tsv")

    def test_detect_headers_csv_pipe_sniff(self, test_db):
        from app.utils.tabular_parser import detect_headers

        columns, delimiter = detect_headers(b"code|name|city\n", "vendors.csv")
        assert columns == ["code", "name", "city"]
        assert delimiter == "|"

    def test_detect_headers_xlsx_delimiter_is_none(self, test_db):
        from app.utils.tabular_parser import detect_headers

        content = _make_xlsx([["code", "name", "city"], ["001", "Acme", "Paris"]])
        columns, delimiter = detect_headers(content, "vendors.xlsx")
        assert columns == ["code", "name", "city"]
        assert delimiter is None

    def test_detect_headers_unknown_extension_raises(self, test_db):
        import pytest

        from app.utils.tabular_parser import detect_headers

        with pytest.raises(ValueError, match="Unsupported file format"):
            detect_headers(b"junk", "vendors.pdf")


def test_parse_file_rejects_tsv(test_db):
    """Core ingestion parsing accepts only CSV and XLSX files."""
    import pytest

    from app.utils.tabular_parser import parse_file

    with pytest.raises(ValueError, match="Unsupported file format"):
        parse_file(b"code\tname\n001\tAcme\n", "vendors.tsv", delimiter="\t")
