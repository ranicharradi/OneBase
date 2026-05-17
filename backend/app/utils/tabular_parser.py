"""Tabular file parsing utilities — dispatches CSV and XLSX inputs to the right reader.

CSV path uses the same ingestion/upload contract:
- UTF-8 with optional BOM (utf-8-sig) required.
- Default delimiter ";" (overridable per data source).
- All values are strings, trimmed of surrounding whitespace.
"""

import csv
import datetime as _dt
import io
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.utils.file_format import extension_of as _extension


def _decode_csv_text(file_content: bytes) -> str:
    """Decode CSV bytes using the only accepted text encoding."""
    try:
        return file_content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV file is not valid UTF-8") from exc


def parse_csv(file_content: bytes, delimiter: str = ";") -> list[dict[str, Any]]:
    """Parse CSV bytes into list of dicts with trimmed values.

    Handles:
    - UTF-8 with optional BOM (utf-8-sig)
    - Semicolon delimiter by default
    - Whitespace trimming on all values
    - Quoted fields with internal delimiters
    """
    if not file_content:
        return []

    text = _decode_csv_text(file_content)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter, quotechar='"')
    rows: list[dict[str, Any]] = []
    for row in reader:
        trimmed = {k.strip(): v.strip() if v else v for k, v in row.items()}
        rows.append(trimmed)
    return rows


def detect_columns_csv(file_content: bytes, delimiter: str = ";") -> list[str]:
    """Extract column headers from the first row of a CSV file."""
    if not file_content:
        return []

    text = _decode_csv_text(file_content)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter, quotechar='"')
    try:
        headers = next(reader)
        return [h.strip() for h in headers]
    except StopIteration:
        return []


def _format_cell(value: Any) -> Any:
    """Normalize an openpyxl cell value for the ingestion pipeline.

    - datetime with zero time → ISO date string (YYYY-MM-DD)
    - datetime with non-zero time → ISO datetime string (YYYY-MM-DDTHH:MM:SS)
    - str → stripped (empty string becomes None)
    - everything else (int, float, bool, None) → unchanged
    """
    if isinstance(value, _dt.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.replace(microsecond=0).isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _load_xlsx(file_content: bytes):
    """Open an xlsx workbook from bytes in read-only mode, raising ValueError on bad input.

    openpyxl can raise InvalidFileException (non-OOXML data),
    BadZipFile (truncated/garbage/empty bytes), or KeyError (missing xl/workbook.xml).
    """
    try:
        return openpyxl.load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, BadZipFile, KeyError, OSError) as exc:
        raise ValueError(f"Could not read Excel file: {exc}") from exc


def parse_xlsx(file_content: bytes) -> list[dict[str, Any]]:
    """Parse the first sheet of an xlsx workbook into a list of dicts.

    Types are preserved (int, float, bool); datetimes become ISO strings;
    strings are trimmed; empty/whitespace strings become None. Trailing
    all-None rows are dropped.
    """
    wb = _load_xlsx(file_content)
    try:
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers: list[str | None] = [
            (h.strip() if isinstance(h, str) else None) if h is not None else None for h in header_row
        ]
        kept_idx = [i for i, h in enumerate(headers) if isinstance(h, str) and h]
        if not kept_idx:
            return []

        kept_headers = [headers[i] for i in kept_idx]

        out: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            if all(cell is None for cell in raw_row):
                continue
            row_dict: dict[str, Any] = {}
            for col_idx, header in zip(kept_idx, kept_headers, strict=False):
                cell = raw_row[col_idx] if col_idx < len(raw_row) else None
                row_dict[header] = _format_cell(cell)
            if any(v is not None for v in row_dict.values()):
                out.append(row_dict)
        return out
    finally:
        wb.close()


def detect_columns_xlsx(file_content: bytes) -> list[str]:
    """Return the trimmed, non-empty headers from the first row of an xlsx workbook."""
    wb = _load_xlsx(file_content)
    try:
        ws = wb.active
        if ws is None:
            return []
        try:
            header_row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            return []
        return [h.strip() for h in header_row if isinstance(h, str) and h.strip()]
    finally:
        wb.close()


_CSV_DELIMITER_CANDIDATES = (",", ";", "\t", "|")

_CSV_EXTENSIONS = {".csv"}
_XLSX_EXTENSIONS = {".xlsx"}


def _sniff_csv_delimiter(file_content: bytes) -> str:
    """Pick the delimiter from , ; \\t | that yields the most header tokens.

    Mirrors the previous front-end heuristic in csvHeaders.ts. Defaults to ';'
    on empty input to match the historical CSV default.
    """
    if not file_content:
        return ";"
    text = _decode_csv_text(file_content[:65536])

    first_line = next(
        (line for line in text.splitlines() if line.strip()),
        "",
    )
    if not first_line:
        return ";"

    best = _CSV_DELIMITER_CANDIDATES[0]
    best_count = -1
    for candidate in _CSV_DELIMITER_CANDIDATES:
        reader = csv.reader(io.StringIO(first_line), delimiter=candidate, quotechar='"')
        try:
            count = len(next(reader))
        except StopIteration:
            count = 0
        if count > best_count:
            best = candidate
            best_count = count
    return best


def parse_file(file_content: bytes, filename: str, delimiter: str = ";") -> list[dict[str, Any]]:
    """Parse a tabular file by extension. delimiter is used for CSV only."""
    ext = _extension(filename)
    if ext in _CSV_EXTENSIONS:
        return parse_csv(file_content, delimiter=delimiter)
    if ext in _XLSX_EXTENSIONS:
        return parse_xlsx(file_content)
    raise ValueError(f"Unsupported file format: {ext or filename!r}")


def detect_headers(file_content: bytes, filename: str) -> tuple[list[str], str | None]:
    """Return (columns, effective_delimiter) for a tabular file.

    For CSV the delimiter is sniffed from `,;\\t|` (most-tokens wins).
    For XLSX the second value is None.
    """
    ext = _extension(filename)
    if ext in _CSV_EXTENSIONS:
        delimiter = _sniff_csv_delimiter(file_content)
        return detect_columns_csv(file_content, delimiter=delimiter), delimiter
    if ext in _XLSX_EXTENSIONS:
        return detect_columns_xlsx(file_content), None
    raise ValueError(f"Unsupported file format: {ext or filename!r}")
